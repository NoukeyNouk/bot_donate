from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import Message, FSInputFile
from aiogram.fsm.storage.memory import MemoryStorage
import asyncio
from markups import *
from openpyxl import *
import random

# Инициализация бота и диспетчера
bot = Bot(token="7992135782:AAFliEUv_5ctfv0iSidq4vIL3nLtEP3WLSI")
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
file_ids = [f'README_md/read{i}.jpg' for i in range(1, 13)]
def from_frase_to_number(word):
    word = list(word)
    word = list(filter(lambda x: x in '1234567890', word))
    word = '+' + ''.join(word)
    return word

class Auth(StatesGroup):
    waiting_for_phone = State()
    confirming_phone = State()
    waiting_for_fio = State()
    confirming_fio = State()

#class Authtorized(StatesGroup):


# Обработчик команды /start
@dp.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    await message.answer_photo(photo=FSInputFile(random.choice(file_ids)), caption="Привет! это вступительное письмою. Для аторизации/регистрации в системе отправьте свой номер телефона")
    await state.set_state(Auth.waiting_for_phone)



@dp.message(Auth.waiting_for_phone)
async def process_name(message: types.Message, state: FSMContext):

    await state.update_data(phone=from_frase_to_number(message.text))  # Сохраняем имя
    user_data = await state.get_data()

    await message.answer_photo(photo=FSInputFile(random.choice(file_ids)), caption=f'Проверьте корректность номера телефона:\n{user_data['phone']}',reply_markup=keyboard)




@dp.callback_query()
async def allbak_handle(callback: types.CallbackQuery, state: FSMContext):
    data = callback.data
    if data == "retry_waiting_for_phone":
        await state.set_state(Auth.waiting_for_phone)
        await callback.message.edit_caption(caption="Введите номер телефона")
    elif data == 'authtorization':
        user_data = await state.get_data()
        await callback.message.edit_caption(caption=f'Проверьте корректность номера телефона:\n{user_data['phone']}')

        await state.set_state(Auth.confirming_phone)
        workbook = load_workbook("База ДД (1).xlsx")
        sheet = workbook.active
        non_empty_rows = 0
        print(1)
        for row in sheet.iter_rows(values_only=True):  # values_only=True возвращает значения, а не объекты ячеек
            if any(cell is not None for cell in row):  # Проверяем, есть ли в строке непустые ячейки
                non_empty_rows += 1
        print(2)
        for i in range(non_empty_rows):

            cell_value = sheet['I'+str(i+2)].value
            print(cell_value)
            fl=0
            if cell_value == user_data['phone']:
                fl=1
                await callback.message.answer_photo(photo=FSInputFile(random.choice(file_ids)), caption=f'Это Ваше ФИО?\n{sheet['A'+str(i+2)].value}',reply_markup=keyboard1)
                await state.set_state(Auth.waiting_for_fio)


        if fl==0:
            await callback.message.answer_photo(photo=FSInputFile(random.choice(file_ids)), caption=f'Введите Ваше ФИО')
            await state.set_state(Auth.waiting_for_fio)

    elif data == "retry_waiting_for_fio":
        await state.set_state(Auth.waiting_for_fio)
        await callback.message.edit_caption(caption="Введите ваше ФИО")
    elif data == 'podtverzdenie':
        user_data = await state.get_data()
        await callback.message.edit_caption(caption=f'Проверьте корректность Вашего ФИО:\n{user_data['phone']}')

        #await state.set_state(Auth.confirming_fio)
        workbook = load_workbook("База ДД (1).xlsx")
        sheet = workbook.active


        print(1,2,3,4)
        user_data = await state.get_data()





@dp.message(Auth.waiting_for_fio)
async def process_name(message: types.Message, state: FSMContext):
    print('erter')
    await state.update_data(phone=message.text)  # Сохраняем имя
    data = await state.get_data()

    await message.answer_photo(photo=FSInputFile(random.choice(file_ids)), caption=f'Проверьте корректность Вашего ФИО:\n{data['phone']}',reply_markup=keyboard1)



# Запуск бота
async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())